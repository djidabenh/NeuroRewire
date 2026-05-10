"""
NeuroReWire — Export Firebase → Excel
=====================================
Exporte toute la base de données Firebase en un fichier Excel structuré.

Structure du fichier généré :
  - Feuille 1 : Utilisateurs       (users/{uid})
  - Feuille 2 : Séances            (sessions/{uid}/{sessionId})
  - Feuille 3 : Résultats exercices (exercise_results/{uid}/{resultId})
  - Feuille 4 : Résumé par catégorie (statistiques agrégées par utilisateur)

Usage :
  python export_to_excel.py

Prérequis :
  pip install firebase-admin openpyxl
"""

import json
import os
import sys
from datetime import datetime

# ── Dépendances ────────────────────────────────────────────────────────────────
try:
    import firebase_admin
    from firebase_admin import credentials, db as firebase_db
except ImportError:
    sys.exit("❌  Installe firebase-admin :  pip install firebase-admin")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("❌  Installe openpyxl :  pip install openpyxl")


# ── Configuration ──────────────────────────────────────────────────────────────
# Option 1 : fichier service account JSON
SERVICE_ACCOUNT_PATH = os.getenv(
    "FIREBASE_SERVICE_ACCOUNT_PATH",
    "./secrets/firebase-service-account.json"
)
# Option 2 : contenu JSON dans une variable d'environnement
SERVICE_ACCOUNT_JSON = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON", "")

DATABASE_URL = os.getenv(
    "FIREBASE_DATABASE_URL",
    "https://neuro-rewire-default-rtdb.firebaseio.com"
)

OUTPUT_FILE = f"neurorewire_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


# ── Palette de couleurs ────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1A5276"   # bleu foncé
CLR_HEADER_FG   = "FFFFFF"   # blanc
CLR_SECTION_BG  = "D6EAF8"   # bleu clair (ligne titre de section)
CLR_ROW_ALT     = "EBF5FB"   # alternance de lignes
CLR_ACCENT      = "2E86C1"   # bleu accent (résumé)
CLR_GREEN       = "1E8449"
CLR_ORANGE      = "CA6F1E"
CLR_RED         = "922B21"


# ── Helpers de style ──────────────────────────────────────────────────────────
def _header_font():
    return Font(name="Arial", bold=True, color=CLR_HEADER_FG, size=10)

def _body_font(bold=False):
    return Font(name="Arial", bold=bold, size=10)

def _header_fill():
    return PatternFill("solid", fgColor=CLR_HEADER_BG)

def _alt_fill():
    return PatternFill("solid", fgColor=CLR_ROW_ALT)

def _thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font   = _header_font()
        c.fill   = _header_fill()
        c.border = _thin_border()
        c.alignment = _center()

def style_data_row(ws, row, ncols, alt=False):
    fill = _alt_fill() if alt else None
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font   = _body_font()
        c.border = _thin_border()
        c.alignment = _left()
        if fill:
            c.fill = fill

def autofit(ws, min_w=10, max_w=45):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + 2, min_w), max_w)

def fmt_date(iso):
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(iso)

def fmt_duration(secs):
    if secs is None:
        return ""
    try:
        secs = int(secs)
        h, r = divmod(secs, 3600)
        m, s = divmod(r, 60)
        if h:
            return f"{h}h {m:02d}m {s:02d}s"
        return f"{m}m {s:02d}s"
    except Exception:
        return str(secs)

def pct(score, max_score):
    try:
        if max_score and int(max_score) > 0:
            return round(int(score) / int(max_score) * 100, 1)
    except Exception:
        pass
    return None


# ── Connexion Firebase ─────────────────────────────────────────────────────────
def init_firebase():
    if SERVICE_ACCOUNT_JSON:
        cred_dict = json.loads(SERVICE_ACCOUNT_JSON)
        cred = credentials.Certificate(cred_dict)
    elif os.path.exists(SERVICE_ACCOUNT_PATH):
        cred = credentials.Certificate(SERVICE_ACCOUNT_PATH)
    else:
        sys.exit(
            "❌  Credentials Firebase introuvables.\n"
            "    → Définis FIREBASE_SERVICE_ACCOUNT_PATH ou FIREBASE_SERVICE_ACCOUNT_JSON"
        )
    firebase_admin.initialize_app(cred, {"databaseURL": DATABASE_URL})
    print(f"✅  Connecté à {DATABASE_URL}")


def fetch_all():
    print("📥  Téléchargement des données…")
    users_snap    = firebase_db.reference("users").get() or {}
    sessions_snap = firebase_db.reference("sessions").get() or {}
    results_snap  = firebase_db.reference("exercise_results").get() or {}
    print(f"   → {len(users_snap)} utilisateur(s)")
    print(f"   → sessions pour {len(sessions_snap)} uid(s)")
    print(f"   → résultats pour {len(results_snap)} uid(s)")
    return users_snap, sessions_snap, results_snap


# ── Feuille 1 : Utilisateurs ──────────────────────────────────────────────────
def sheet_users(wb, users):
    ws = wb.create_sheet("Utilisateurs")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = [
        "UID", "Prénom", "Nom", "Date naissance", "Email",
        "Wilaya", "Téléphone", "Hôpital",
        "Type score AVC", "Valeur score AVC",
        "Date inscription", "Dernière modification"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    for i, (uid, u) in enumerate(users.items(), start=2):
        row = [
            uid,
            u.get("givenName", ""),
            u.get("familyName", ""),
            fmt_date(u.get("birthDate", "")),
            u.get("email", ""),
            u.get("wilaya", ""),
            u.get("phone", ""),
            u.get("hospital", ""),
            u.get("strokeScoreType", ""),
            u.get("strokeScoreValue", ""),
            fmt_date(u.get("createdAt", "")),
            fmt_date(u.get("updatedAt", "")),
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=i, column=col, value=val)
        style_data_row(ws, i, len(headers), alt=(i % 2 == 0))

    # Ligne totaux
    total_row = i + 2
    ws.cell(row=total_row, column=1, value=f"Total : {len(users)} utilisateur(s)")
    ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, size=10)

    autofit(ws)
    return ws


# ── Feuille 2 : Séances ───────────────────────────────────────────────────────
def sheet_sessions(wb, users, sessions_by_uid):
    ws = wb.create_sheet("Séances")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = [
        "UID", "Prénom", "Nom", "Session ID",
        "Début", "Fin", "Durée", "Statut", "Source", "Notes",
        "Créée le", "Modifiée le"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    row_i = 2
    total_sessions = 0
    for uid, sessions in sessions_by_uid.items():
        u = users.get(uid, {})
        fname = u.get("givenName", "")
        lname = u.get("familyName", "")
        for sid, s in sessions.items():
            total_sessions += 1
            status = s.get("status", "")
            status_label = {
                "completed": "✅ Terminée",
                "started":   "🔄 En cours",
                "cancelled": "❌ Annulée",
            }.get(status, status)
            row = [
                uid, fname, lname, sid,
                fmt_date(s.get("startedAt")),
                fmt_date(s.get("endedAt")),
                fmt_duration(s.get("durationSeconds")),
                status_label,
                s.get("source", ""),
                s.get("notes", ""),
                fmt_date(s.get("createdAt")),
                fmt_date(s.get("updatedAt")),
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=row_i, column=col, value=val)
            style_data_row(ws, row_i, len(headers), alt=(row_i % 2 == 0))
            row_i += 1

    total_row = row_i + 1
    ws.cell(row=total_row, column=1, value=f"Total : {total_sessions} séance(s)")
    ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, size=10)

    autofit(ws)
    return ws


# ── Feuille 3 : Résultats d'exercices ─────────────────────────────────────────
def sheet_results(wb, users, results_by_uid):
    ws = wb.create_sheet("Résultats exercices")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = [
        "UID", "Prénom", "Nom", "Result ID", "Session ID",
        "Catégorie", "Exercice (clé)",
        "Début", "Fin", "Durée",
        "Score", "Score max", "% Réussite",
        "Statut", "Créé le", "Modifié le"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    row_i = 2
    total_results = 0
    for uid, results in results_by_uid.items():
        u = users.get(uid, {})
        fname = u.get("givenName", "")
        lname = u.get("familyName", "")
        for rid, r in results.items():
            total_results += 1
            score     = r.get("score")
            max_score = r.get("maxScore")
            p         = pct(score, max_score)
            status    = r.get("status", "")
            status_label = {
                "completed": "✅ Terminé",
                "started":   "🔄 En cours",
                "failed":    "❌ Échoué",
                "abandoned": "⏹ Abandonné",
            }.get(status, status)

            row = [
                uid, fname, lname, rid,
                r.get("sessionId", ""),
                r.get("category", ""),
                r.get("exerciseKey", ""),
                fmt_date(r.get("startedAt")),
                fmt_date(r.get("endedAt")),
                fmt_duration(r.get("durationSeconds")),
                score if score is not None else "",
                max_score if max_score is not None else "",
                f"{p}%" if p is not None else "",
                status_label,
                fmt_date(r.get("createdAt")),
                fmt_date(r.get("updatedAt")),
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=row_i, column=col, value=val)
            style_data_row(ws, row_i, len(headers), alt=(row_i % 2 == 0))

            # Couleur % réussite
            pct_cell = ws.cell(row=row_i, column=13)
            if p is not None:
                if p >= 80:
                    pct_cell.font = Font(name="Arial", bold=True, color=CLR_GREEN, size=10)
                elif p >= 50:
                    pct_cell.font = Font(name="Arial", bold=True, color=CLR_ORANGE, size=10)
                else:
                    pct_cell.font = Font(name="Arial", bold=True, color=CLR_RED, size=10)

            row_i += 1

    total_row = row_i + 1
    ws.cell(row=total_row, column=1, value=f"Total : {total_results} résultat(s)")
    ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, size=10)

    autofit(ws)
    return ws


# ── Feuille 4 : Résumé par utilisateur ────────────────────────────────────────
CATEGORIES = ["memory", "attention", "coordination", "trajectory", "language", "quiz"]

def sheet_summary(wb, users, sessions_by_uid, results_by_uid):
    ws = wb.create_sheet("Résumé utilisateurs")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40

    cat_headers = []
    for cat in CATEGORIES:
        cat_headers += [f"{cat}\nJeux", f"{cat}\n% moy"]

    headers = (
        ["UID", "Prénom", "Nom", "Email", "Wilaya", "Hôpital",
         "Score AVC", "Inscription",
         "Séances total", "Séances terminées",
         "Exercices total", "Exercices terminés",
         "Temps total"]
        + cat_headers
    )
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = _header_font()
        c.fill      = _header_fill()
        c.border    = _thin_border()
        c.alignment = _center()

    for i, (uid, u) in enumerate(users.items(), start=2):
        sessions = sessions_by_uid.get(uid, {})
        results  = results_by_uid.get(uid, {})

        total_sessions     = len(sessions)
        completed_sessions = sum(1 for s in sessions.values() if s.get("status") == "completed")

        total_results     = len(results)
        completed_results = [r for r in results.values() if r.get("status") == "completed"]
        total_secs        = sum(
            int(r.get("durationSeconds", 0) or 0) for r in completed_results
        )

        # Stats par catégorie
        cat_stats = {}
        for cat in CATEGORIES:
            cat_results = [r for r in completed_results if r.get("category") == cat]
            count = len(cat_results)
            scores = [pct(r.get("score"), r.get("maxScore")) for r in cat_results]
            scores = [s for s in scores if s is not None]
            avg   = round(sum(scores) / len(scores), 1) if scores else None
            cat_stats[cat] = (count, avg)

        stroke_score = ""
        if u.get("strokeScoreType") and u.get("strokeScoreValue"):
            stroke_score = f"{u['strokeScoreType']} : {u['strokeScoreValue']}"

        row_vals = [
            uid,
            u.get("givenName", ""),
            u.get("familyName", ""),
            u.get("email", ""),
            u.get("wilaya", ""),
            u.get("hospital", ""),
            stroke_score,
            fmt_date(u.get("createdAt", "")),
            total_sessions,
            completed_sessions,
            total_results,
            len(completed_results),
            fmt_duration(total_secs),
        ]
        for cat in CATEGORIES:
            count, avg = cat_stats[cat]
            row_vals.append(count if count else "")
            row_vals.append(f"{avg}%" if avg is not None else "")

        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font      = _body_font()
            c.border    = _thin_border()
            c.alignment = _left()
            if i % 2 == 0:
                c.fill = _alt_fill()

    autofit(ws)
    return ws


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    init_firebase()
    users, sessions_by_uid, results_by_uid = fetch_all()

    print("📊  Génération du fichier Excel…")
    wb = Workbook()
    wb.remove(wb.active)  # retire la feuille vide par défaut

    sheet_users(wb, users)
    sheet_sessions(wb, users, sessions_by_uid)
    sheet_results(wb, users, results_by_uid)
    sheet_summary(wb, users, sessions_by_uid, results_by_uid)

    wb.save(OUTPUT_FILE)
    print(f"✅  Fichier généré : {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
