"""
NeuroReWire — Export Dataset plat pour entraînement ML
=======================================================
Génère un fichier Excel avec UNE LIGNE PAR EXERCICE TERMINÉ,
toutes les colonnes à plat (profil utilisateur + résultat + metadata du jeu).

Colonnes produites :
  — Profil utilisateur (8 colonnes)
  — Session (3 colonnes)
  — Résultat de base (10 colonnes)
  — Metadata communes (2 colonnes)
  — Metadata spécifiques par jeu (18 colonnes, NaN si non applicable)

Usage :
  python export_dataset.py

Prérequis :
  pip install firebase-admin openpyxl
"""

import json
import os
import sys
from datetime import datetime

try:
    import firebase_admin
    from firebase_admin import credentials, db as firebase_db
except ImportError:
    sys.exit("❌  pip install firebase-admin")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("❌  pip install openpyxl")


# ── Config ─────────────────────────────────────────────────────────────────────
# Chemin vers ton fichier service account Firebase (déjà dans ton projet)
SERVICE_ACCOUNT_PATH = os.getenv("FIREBASE_SERVICE_ACCOUNT_PATH",
                                  "./secrets/firebase-service-account.json")
SERVICE_ACCOUNT_JSON = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON", "")
# DATABASE_URL est lue automatiquement depuis le service account si non définie
DATABASE_URL = os.getenv("FIREBASE_DATABASE_URL", "")
OUTPUT_FILE = f"neurorewire_dataset_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


# ── Colonnes du dataset ────────────────────────────────────────────────────────
# Chaque entrée : (nom_colonne, description)
COLUMNS = [
    # ── Identifiants ──────────────────────────────────────────────────────────
    ("uid",                     "Identifiant unique Firebase de l'utilisateur"),
    ("result_id",               "Identifiant unique du résultat d'exercice"),
    ("session_id",              "Identifiant de la séance"),

    # ── Profil utilisateur ────────────────────────────────────────────────────
    ("given_name",              "Prénom"),
    ("family_name",             "Nom de famille"),
    ("birth_date",              "Date de naissance (brute)"),
    ("age_at_exercise",         "Âge calculé au moment de l'exercice (années)"),
    ("wilaya",                  "Wilaya / région"),
    ("hospital",                "Établissement de santé"),
    ("stroke_score_type",       "Type de score AVC (ex: NIHSS, Rankin)"),
    ("stroke_score_value",      "Valeur du score AVC"),

    # ── Session ───────────────────────────────────────────────────────────────
    ("session_started_at",      "Début de la séance (ISO)"),
    ("session_status",          "Statut de la séance (completed/started/cancelled)"),
    ("session_duration_sec",    "Durée totale de la séance (secondes)"),

    # ── Résultat de base ──────────────────────────────────────────────────────
    ("category",                "Catégorie cognitive (memory/attention/coordination/trajectory/language/quiz)"),
    ("exercise_key",            "Clé unique de l'exercice (ex: memory-images)"),
    ("status",                  "Statut de l'exercice (completed/failed/abandoned)"),
    ("started_at",              "Début de l'exercice (ISO)"),
    ("ended_at",                "Fin de l'exercice (ISO)"),
    ("duration_sec",            "Durée de l'exercice (secondes)"),
    ("score",                   "Score brut obtenu"),
    ("max_score",               "Score maximum possible"),
    ("score_pct",               "Pourcentage de réussite (0–100)"),
    ("created_at",              "Timestamp de création du résultat (ISO)"),

    # ── Metadata communes ─────────────────────────────────────────────────────
    ("rounds",                  "Nombre de rounds / niveaux joués"),
    ("accuracy_pct",            "Précision calculée par le jeu (0–100)"),

    # ── Metadata : Attention — Cliquer cible (attention-click-target) ─────────
    ("targets_shown",           "[attention-click] Cibles affichées"),
    ("targets_hit",             "[attention-click] Cibles touchées"),
    ("wrong_clicks",            "[attention-click/scene] Clics incorrects"),

    # ── Metadata : Attention — Scène (attention-scene-search) ─────────────────
    ("targets_found",           "[scene-search] Cibles trouvées"),

    # ── Metadata : Attention — Intrus (attention-intruder) ────────────────────
    ("correct_selections",      "[intruder] Sélections correctes"),

    # ── Metadata : Coordination — Attraper fruits (coordination-catch) ────────
    ("fruits_shown",            "[catch] Fruits apparus"),
    ("fruits_caught",           "[catch] Fruits attrapés"),
    ("fruits_missed",           "[catch] Fruits ratés"),

    # ── Metadata : Coordination — Drag & Drop ────────────────────────────────
    ("correct_drops",           "[drag-drop] Dépôts corrects"),
    ("wrong_drops",             "[drag-drop] Dépôts incorrects"),
    ("total_objects",           "[drag-drop] Total objets"),
    ("total_placed_objects",    "[drag-drop] Total objets placés"),

    # ── Metadata : Coordination — Tap séquence ───────────────────────────────
    ("correct_hits",            "[tap] Touches correctes"),
    ("wrong_hits",              "[tap] Touches incorrectes"),

    # ── Metadata : Mémoire ────────────────────────────────────────────────────
    ("pairs_found",             "[pairs] Paires trouvées"),
    ("attempts",                "[pairs] Tentatives totales"),
    ("mistakes",                "[pairs] Erreurs"),
    ("correct_placements",      "[order] Placements corrects"),
    ("total_placements",        "[order] Total placements"),

    # ── Metadata : Trajectoire ────────────────────────────────────────────────
    ("segments_completed",      "[dots] Segments complétés"),
    ("mazes_completed",         "[maze] Labyrinthes complétés"),
    ("wall_hits",               "[maze] Touches de mur"),
    ("successful_shapes",       "[shapes] Formes réussies"),
    ("traj_errors",             "[traj] Erreurs (dots/shapes)"),

    # ── Metadata : Langage ────────────────────────────────────────────────────
    ("items_shown",             "[language] Items présentés"),
    ("correct_pronunciations",  "[language] Prononciations correctes"),
    ("wrong_pronunciations",    "[language] Prononciations incorrectes"),
    ("speech_recognition",      "[language] Reconnaissance vocale activée (0/1)"),
    ("language_mode",           "[language] Mode du jeu (pronunciation-scored, etc.)"),

    # ── Metadata : Quiz ───────────────────────────────────────────────────────
    ("questions_shown",         "[quiz] Questions / comparaisons affichées"),
    ("correct_answers",         "[quiz] Réponses correctes"),
    ("wrong_answers",           "[quiz] Réponses incorrectes"),
    ("timed_mode",              "[quiz] Mode chronométré (0/1)"),
]

COL_NAMES = [c[0] for c in COLUMNS]
COL_DESC  = [c[1] for c in COLUMNS]


# ── Firebase ───────────────────────────────────────────────────────────────────
def init_firebase():
    global DATABASE_URL

    # ── Charger les credentials ────────────────────────────────────────────────
    if SERVICE_ACCOUNT_JSON:
        cred_dict = json.loads(SERVICE_ACCOUNT_JSON)
        cred = credentials.Certificate(cred_dict)
    elif os.path.exists(SERVICE_ACCOUNT_PATH):
        with open(SERVICE_ACCOUNT_PATH, "r", encoding="utf-8") as f:
            cred_dict = json.load(f)
        cred = credentials.Certificate(SERVICE_ACCOUNT_PATH)
    else:
        print(f"\n❌  Fichier service account introuvable : {SERVICE_ACCOUNT_PATH}")
        print("    Assure-toi que le fichier secrets/firebase-service-account.json")
        print("    existe dans le dossier neurorewire/")
        sys.exit(1)

    # ── Déduire DATABASE_URL depuis le project_id si non définie ──────────────
    if not DATABASE_URL:
        project_id = cred_dict.get("project_id", "")
        if project_id:
            DATABASE_URL = f"https://{project_id}-default-rtdb.firebaseio.com"
            print(f"ℹ️   DATABASE_URL déduite : {DATABASE_URL}")
        else:
            print("\n❌  Impossible de déduire DATABASE_URL.")
            print("    Définis la variable FIREBASE_DATABASE_URL ou vérifie ton service account.")
            sys.exit(1)

    firebase_admin.initialize_app(cred, {"databaseURL": DATABASE_URL})
    print(f"✅  Connecté à Firebase : {DATABASE_URL}")


def fetch_all():
    print("📥  Téléchargement des données depuis Firebase…")

    try:
        users = firebase_db.reference("users").get() or {}
    except Exception as e:
        print(f"\n❌  Erreur lecture 'users' : {e}")
        print("    Vérifie que DATABASE_URL est correct et que les règles Firebase")
        print("    autorisent la lecture depuis le service account.")
        sys.exit(1)

    sessions = firebase_db.reference("sessions").get() or {}
    results  = firebase_db.reference("exercise_results").get() or {}

    n_sessions = sum(len(v) for v in sessions.values()) if sessions else 0
    n_results  = sum(len(v) for v in results.values()) if results else 0

    print(f"   ✔ {len(users)} utilisateur(s)")
    print(f"   ✔ {n_sessions} séance(s)")
    print(f"   ✔ {n_results} résultat(s) d'exercice")

    if not users and not results:
        print("\n⚠️   La base de données semble vide.")
        print("    Vérifie :")
        print("    1. Que DATABASE_URL pointe vers la bonne base Firebase")
        print("    2. Que les règles Firebase permettent la lecture avec ce service account")
        print("    3. Que des données existent dans Firebase Realtime Database")

    return users, sessions, results


# ── Helpers ────────────────────────────────────────────────────────────────────
def calc_age(birth_date_str, reference_iso):
    try:
        bd = datetime.fromisoformat(birth_date_str.replace("Z", "+00:00"))
        ref = datetime.fromisoformat(reference_iso.replace("Z", "+00:00"))
        return ref.year - bd.year - ((ref.month, ref.day) < (bd.month, bd.day))
    except Exception:
        return None

def safe_pct(score, max_score):
    try:
        s, m = float(score), float(max_score)
        return round(s / m * 100, 2) if m > 0 else None
    except Exception:
        return None

def g(d, *keys, default=None):
    """Accès sécurisé imbriqué dans un dict."""
    for k in keys:
        if not isinstance(d, dict):
            return default
        d = d.get(k, default)
        if d is None:
            return default
    return d

def bool_int(val):
    if val is None:
        return None
    return 1 if val else 0


# ── Construction d'une ligne dataset ──────────────────────────────────────────
def build_row(uid, result_id, result, user, session):
    m = result.get("metadata") or {}
    started_at = result.get("startedAt", "")

    row = {col: None for col in COL_NAMES}

    # Identifiants
    row["uid"]       = uid
    row["result_id"] = result_id
    row["session_id"] = result.get("sessionId", "")

    # Profil
    row["given_name"]         = user.get("givenName", "")
    row["family_name"]        = user.get("familyName", "")
    row["birth_date"]         = user.get("birthDate", "")
    row["age_at_exercise"]    = calc_age(user.get("birthDate", ""), started_at) if started_at else None
    row["wilaya"]             = user.get("wilaya", "")
    row["hospital"]           = user.get("hospital", "")
    row["stroke_score_type"]  = user.get("strokeScoreType", "")
    row["stroke_score_value"] = user.get("strokeScoreValue", "")

    # Session
    if session:
        row["session_started_at"]   = session.get("startedAt", "")
        row["session_status"]       = session.get("status", "")
        row["session_duration_sec"] = session.get("durationSeconds")

    # Résultat de base
    row["category"]     = result.get("category", "")
    row["exercise_key"] = result.get("exerciseKey", "")
    row["status"]       = result.get("status", "")
    row["started_at"]   = result.get("startedAt", "")
    row["ended_at"]     = result.get("endedAt", "")
    row["duration_sec"] = result.get("durationSeconds")
    row["score"]        = result.get("score")
    row["max_score"]    = result.get("maxScore")
    row["score_pct"]    = safe_pct(result.get("score"), result.get("maxScore"))
    row["created_at"]   = result.get("createdAt", "")

    # Metadata communes
    row["rounds"]       = m.get("rounds")
    row["accuracy_pct"] = m.get("accuracyPercent")

    # Attention — click target
    row["targets_shown"] = m.get("targetsShown")
    row["targets_hit"]   = m.get("targetsHit")
    row["wrong_clicks"]  = m.get("wrongClicks")

    # Attention — scene search
    row["targets_found"] = m.get("targetsFound")

    # Attention — intruder
    row["correct_selections"] = m.get("correctSelections")

    # Coordination — catch
    row["fruits_shown"]  = m.get("fruitsShown")
    row["fruits_caught"] = m.get("fruitsCaught")
    row["fruits_missed"] = m.get("fruitsMissed")

    # Coordination — drag-drop
    row["correct_drops"]       = m.get("correctDrops")
    row["wrong_drops"]         = m.get("wrongDrops")
    row["total_objects"]       = m.get("totalObjects")
    row["total_placed_objects"]= m.get("totalPlacedObjects")

    # Coordination — tap
    row["correct_hits"] = m.get("correctHits")
    row["wrong_hits"]   = m.get("wrongHits")

    # Mémoire — pairs
    row["pairs_found"] = m.get("pairsFound")
    row["attempts"]    = m.get("attempts")
    row["mistakes"]    = m.get("mistakes")

    # Mémoire — order
    row["correct_placements"] = m.get("correctPlacements")
    row["total_placements"]   = m.get("totalPlacements")

    # Trajectoire
    row["segments_completed"] = m.get("segmentsCompleted")
    row["mazes_completed"]    = m.get("mazesCompleted")
    row["wall_hits"]          = m.get("wallHits")
    row["successful_shapes"]  = m.get("successfulShapes")
    row["traj_errors"]        = m.get("errors")

    # Langage
    row["items_shown"]            = m.get("itemsShown")
    row["correct_pronunciations"] = m.get("correctPronunciations")
    row["wrong_pronunciations"]   = m.get("wrongPronunciations")
    row["speech_recognition"]     = bool_int(m.get("speechRecognitionEnabled"))
    row["language_mode"]          = m.get("mode", "")

    # Quiz
    row["questions_shown"]  = m.get("questionsShown") or m.get("comparisonsShown")
    row["correct_answers"]  = m.get("correctAnswers")
    row["wrong_answers"]    = m.get("wrongAnswers")
    row["timed_mode"]       = bool_int(m.get("timedMode"))

    return [row[col] for col in COL_NAMES]


# ── Styles Excel ───────────────────────────────────────────────────────────────
HEADER_BG  = "1A3A5C"
HEADER_FG  = "FFFFFF"
DESC_BG    = "D0DCF0"
ALT_BG     = "EEF4FB"

def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, nrow, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=nrow, column=col)
        c.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=9)
        c.fill      = PatternFill("solid", fgColor=HEADER_BG)
        c.border    = thin()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_desc(ws, nrow, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=nrow, column=col)
        c.font      = Font(name="Arial", italic=True, color="444444", size=8)
        c.fill      = PatternFill("solid", fgColor=DESC_BG)
        c.border    = thin()
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_data(ws, nrow, ncols, alt):
    fill = PatternFill("solid", fgColor=ALT_BG) if alt else None
    for col in range(1, ncols + 1):
        c = ws.cell(row=nrow, column=col)
        c.font      = Font(name="Arial", size=9)
        c.border    = thin()
        c.alignment = Alignment(horizontal="left", vertical="center")
        if fill:
            c.fill = fill

def autofit(ws, mn=8, mx=35):
    for col_cells in ws.columns:
        w = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, mn), mx)


# ── Génération Excel ───────────────────────────────────────────────────────────
def generate_excel(users, sessions_by_uid, results_by_uid):
    wb = Workbook()
    ncols = len(COL_NAMES)

    # ── Feuille 1 : Dataset ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dataset"
    ws.freeze_panes = "A3"   # fige les 2 lignes d'en-tête
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 48

    # Ligne 1 : noms de colonnes
    for col, name in enumerate(COL_NAMES, 1):
        ws.cell(row=1, column=col, value=name)
    apply_header(ws, 1, ncols)

    # Ligne 2 : descriptions
    for col, desc in enumerate(COL_DESC, 1):
        ws.cell(row=2, column=col, value=desc)
    apply_desc(ws, 2, ncols)

    # Données
    data_row = 3
    total = 0
    for uid, results in results_by_uid.items():
        user     = users.get(uid, {})
        sessions = sessions_by_uid.get(uid, {})
        for result_id, result in results.items():
            # On inclut TOUS les statuts (completed, failed, abandoned, started)
            sid     = result.get("sessionId")
            session = sessions.get(sid) if sid else None
            row_vals = build_row(uid, result_id, result, user, session)
            for col, val in enumerate(row_vals, 1):
                ws.cell(row=data_row, column=col, value=val)
            apply_data(ws, data_row, ncols, alt=(data_row % 2 == 0))
            data_row += 1
            total += 1

    autofit(ws)

    # ── Feuille 2 : Dictionnaire des colonnes ──────────────────────────────────
    wd = wb.create_sheet("Dictionnaire")
    wd.column_dimensions["A"].width = 30
    wd.column_dimensions["B"].width = 60
    wd.column_dimensions["C"].width = 20

    headers = ["Colonne", "Description", "Groupe"]
    for col, h in enumerate(headers, 1):
        c = wd.cell(row=1, column=col, value=h)
        c.font   = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        c.fill   = PatternFill("solid", fgColor=HEADER_BG)
        c.border = thin()
        c.alignment = Alignment(horizontal="center", vertical="center")

    groups = {
        "uid": "Identifiants", "result_id": "Identifiants", "session_id": "Identifiants",
        "given_name": "Profil", "family_name": "Profil", "birth_date": "Profil",
        "age_at_exercise": "Profil", "wilaya": "Profil", "hospital": "Profil",
        "stroke_score_type": "Profil", "stroke_score_value": "Profil",
        "session_started_at": "Session", "session_status": "Session",
        "session_duration_sec": "Session",
        "category": "Résultat", "exercise_key": "Résultat", "status": "Résultat",
        "started_at": "Résultat", "ended_at": "Résultat", "duration_sec": "Résultat",
        "score": "Résultat", "max_score": "Résultat", "score_pct": "Résultat",
        "created_at": "Résultat",
        "rounds": "Metadata commune", "accuracy_pct": "Metadata commune",
    }
    group_colors = {
        "Identifiants":      "1A3A5C",
        "Profil":            "1A5276",
        "Session":           "117A65",
        "Résultat":          "7D6608",
        "Metadata commune":  "6E2F9E",
        "Attention":         "C0392B",
        "Coordination":      "D35400",
        "Mémoire":           "1E8449",
        "Trajectoire":       "2471A3",
        "Langage":           "76448A",
        "Quiz":              "17202A",
    }

    for i, (col_name, desc) in enumerate(COLUMNS, start=2):
        grp = groups.get(col_name)
        if grp is None:
            if col_name.startswith("targets") or col_name in ("wrong_clicks", "targets_found", "correct_selections"):
                grp = "Attention"
            elif col_name.startswith("fruits") or col_name in ("correct_drops", "wrong_drops", "total_objects", "total_placed_objects", "correct_hits", "wrong_hits"):
                grp = "Coordination"
            elif col_name in ("pairs_found", "attempts", "mistakes", "correct_placements", "total_placements"):
                grp = "Mémoire"
            elif col_name in ("segments_completed", "mazes_completed", "wall_hits", "successful_shapes", "traj_errors"):
                grp = "Trajectoire"
            elif col_name in ("items_shown", "correct_pronunciations", "wrong_pronunciations", "speech_recognition", "language_mode"):
                grp = "Langage"
            elif col_name in ("questions_shown", "correct_answers", "wrong_answers", "timed_mode"):
                grp = "Quiz"
            else:
                grp = ""

        color = group_colors.get(grp, "333333")
        for col_i, val in enumerate([col_name, desc, grp], 1):
            c = wd.cell(row=i, column=col_i, value=val)
            c.font   = Font(name="Arial", size=9, bold=(col_i == 1), color=color if col_i != 2 else "000000")
            c.border = thin()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col_i == 2))
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F4F6F7")
        wd.row_dimensions[i].height = 20

    return wb, total


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    init_firebase()
    users, sessions_by_uid, results_by_uid = fetch_all()

    print("📊  Génération du dataset…")
    wb, total = generate_excel(users, sessions_by_uid, results_by_uid)

    wb.save(OUTPUT_FILE)
    print(f"✅  {total} lignes exportées → {OUTPUT_FILE}")
    print(f"   {len(COLUMNS)} colonnes : profil + résultat + toutes les metadata")


if __name__ == "__main__":
    main()
